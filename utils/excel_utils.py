import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from extensions import db
from models import Product, Order, OrderItem, ProductAgency, Agency, Category, UOM, TaxMaster

def export_products_to_excel(products, target_agency_id=None):
    """Export products to Excel file using openpyxl with optional agency filtering"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Headers
    headers = ['ID', 'Name', 'Description', 'SKU', 'Sell Price (Effective)', 'Buy Cost', 'MRP Price (Effective)', 'Category (Effective)', 'GST %', 'Agency', 'Active (Mapping)', 'Created At']
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add data
    for product in products:
        # Find mapping for given agency context
        mapping = None
        
        if target_agency_id:
            # If agency_id is specified, find the specific mapping for that agency
            if hasattr(product, 'agency_mappings') and product.agency_mappings:
                for m in product.agency_mappings:
                    if m.agency_id == target_agency_id:
                        mapping = m
                        break
        else:
            # Otherwise use the first mapping (for all products export)
            if hasattr(product, 'agency_mappings') and product.agency_mappings:
                mapping = product.agency_mappings[0]
        
        # Calculate effective values (prefer agency override, fall back to master)
        effective_sell = float(mapping.sell_price) if mapping and mapping.sell_price is not None else float(product.sell_price or 0)
        effective_buy = float(mapping.buy_price) if mapping and mapping.buy_price is not None else float(product.buy_price or 0)
        effective_mrp = float(mapping.mrp_price) if mapping and mapping.mrp_price is not None else float(product.mrp_price or 0)
        
        effective_category_name = (
            (mapping.category_ref.name if mapping and mapping.category_ref else None)
            or (product.category_ref.name if hasattr(product, 'category_ref') and product.category_ref else None)
            or (product.category if hasattr(product, 'category') else None)
        )
        
        # Get effective tax rate
        effective_tax_rate = ''
        if mapping and mapping.tax_master_ref:
            effective_tax_rate = float(mapping.tax_master_ref.tax_rate) if mapping.tax_master_ref.tax_rate else ''
        elif product.tax_master_ref:
            effective_tax_rate = float(product.tax_master_ref.tax_rate) if product.tax_master_ref.tax_rate else ''
        
        agency_name = mapping.agency.name if mapping else ''
        mapping_active = mapping.is_active if mapping else True
        row_data = [
            product.id,
            product.name,
            product.description,
            product.sku,
            effective_sell,
            effective_buy,
            effective_mrp,
            effective_category_name or '-',
            effective_tax_rate,
            agency_name,
            mapping_active,
            product.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        ws.append(row_data)
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def import_products_from_excel(file, agency_id, user_role):
    """Import products from Excel or CSV file"""
    try:
        imported = 0
        skipped = 0
        
        # Pre-fetch master data for lookups
        categories = {c.name.lower(): c.id for c in Category.query.all()}
        uoms = {u.name.lower(): u.id for u in UOM.query.all()}
        tax_masters = {t.name.lower(): t.id for t in TaxMaster.query.all()}
        
        if file.filename.lower().endswith('.csv'):
            # Handle CSV file
            content = file.read().decode('utf-8')
            csv_file = io.StringIO(content)
            reader = csv.DictReader(csv_file)
            
            for row in reader:
                # Extract data from row
                name = row.get('Name') or row.get('name') or ''
                sku = row.get('SKU') or row.get('sku') or ''
                price = row.get('Price') or row.get('price') or ''
                
                if not all([name.strip(), sku.strip(), price]):
                    skipped += 1
                    continue
                
                # Check if SKU already exists (global)
                if Product.query.filter_by(sku=sku.strip()).first():
                    skipped += 1
                    continue
                
                # Create product master and mapping
                try:
                    # Extract cost, description, and MRP with case-insensitive matching
                    cost_val = row.get('Cost') or row.get('cost') or row.get('COST') or 0
                    desc_val = row.get('Description') or row.get('description') or row.get('DESCRIPTION') or ''
                    mrp_val = row.get('MRP') or row.get('mrp') or row.get('Mrp') or price
                    
                    buy = float(cost_val) if cost_val else 0
                    sell = float(price)
                    mrp = float(mrp_val) if mrp_val else sell
                    margin = round(((sell - buy) / buy) * 100, 2) if buy > 0 else 0

                    # Extract and lookup category, UOM, and tax (check multiple case variants)
                    category_name = str(row.get('Category') or row.get('category') or row.get('CATEGORY') or '').strip()
                    uom_name = str(row.get('UOM') or row.get('uom') or row.get('Uom') or '').strip()
                    tax_name = str(row.get('Tax') or row.get('tax') or row.get('TAX') or '').strip()
                    
                    category_id = categories.get(category_name.lower()) if category_name else None
                    uom_id = uoms.get(uom_name.lower()) if uom_name else None
                    tax_master_id = tax_masters.get(tax_name.lower()) if tax_name else None

                    product = Product(
                        name=name.strip(),
                        description=str(desc_val).strip(),
                        sku=sku.strip(),
                        buy_price=buy,
                        sell_price=sell,
                        mrp_price=mrp,
                        margin=margin,
                        category_id=category_id,
                        uom_id=uom_id,
                        tax_master_id=tax_master_id,
                        is_active=True
                    )
                    db.session.add(product)
                    db.session.flush()

                    mapping = ProductAgency(
                        product_id=product.id,
                        agency_id=agency_id,
                        is_active=True
                    )
                    db.session.add(mapping)
                    imported += 1
                except (ValueError, TypeError):
                    skipped += 1
                    continue
        
        else:
            # Handle Excel file using openpyxl
            from openpyxl import load_workbook
            
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # Process data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                
                # Create row dictionary
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_dict[headers[i]] = value
                
                # Extract data from row
                name = row_dict.get('Name') or row_dict.get('name') or ''
                sku = row_dict.get('SKU') or row_dict.get('sku') or ''
                price = row_dict.get('Price') or row_dict.get('price') or ''
                
                if not all([str(name).strip(), str(sku).strip(), price]):
                    skipped += 1
                    continue
                
                # Check if SKU already exists (global)
                if Product.query.filter_by(sku=str(sku).strip()).first():
                    skipped += 1
                    continue
                
                # Create product master and mapping
                try:
                    # Extract cost, description, and MRP with case-insensitive matching
                    cost_val = row_dict.get('Cost') or row_dict.get('cost') or row_dict.get('COST') or 0
                    desc_val = row_dict.get('Description') or row_dict.get('description') or row_dict.get('DESCRIPTION') or ''
                    mrp_val = row_dict.get('MRP') or row_dict.get('mrp') or row_dict.get('Mrp') or price
                    
                    buy = float(cost_val) if cost_val else 0
                    sell = float(price)
                    mrp = float(mrp_val) if mrp_val else sell
                    margin = round(((sell - buy) / buy) * 100, 2) if buy > 0 else 0

                    # Extract and lookup category, UOM, and tax (check multiple case variants)
                    category_name = str(row_dict.get('Category') or row_dict.get('category') or row_dict.get('CATEGORY') or '').strip()
                    uom_name = str(row_dict.get('UOM') or row_dict.get('uom') or row_dict.get('Uom') or '').strip()
                    tax_name = str(row_dict.get('Tax') or row_dict.get('tax') or row_dict.get('TAX') or '').strip()
                    
                    category_id = categories.get(category_name.lower()) if category_name else None
                    uom_id = uoms.get(uom_name.lower()) if uom_name else None
                    tax_master_id = tax_masters.get(tax_name.lower()) if tax_name else None

                    product = Product(
                        name=str(name).strip(),
                        description=str(desc_val).strip(),
                        sku=str(sku).strip(),
                        buy_price=buy,
                        sell_price=sell,
                        mrp_price=mrp,
                        margin=margin,
                        category_id=category_id,
                        uom_id=uom_id,
                        tax_master_id=tax_master_id,
                        is_active=True
                    )
                    
                    db.session.add(product)
                    db.session.flush()

                    mapping = ProductAgency(
                        product_id=product.id,
                        agency_id=agency_id,
                        is_active=True
                    )
                    db.session.add(mapping)
                    imported += 1
                except (ValueError, TypeError):
                    skipped += 1
                    continue
        
        db.session.commit()
        
        return {
            'success': True,
            'imported': imported,
            'skipped': skipped
        }
        
    except Exception as e:
        db.session.rollback()
        return {
            'success': False,
            'message': str(e)
        }

def export_orders_to_excel(orders):
    """Export orders to Excel file using openpyxl"""
    wb = Workbook()
    
    # Orders details sheet
    ws_details = wb.active
    ws_details.title = "Order Details"
    
    # Headers for details
    detail_headers = [
        'Order ID', 'Order Number', 'Customer', 'Customer Email', 'Customer Phone',
        'Location', 'Agency', 'Salesperson', 'Product Name', 'Product SKU',
        'Quantity', 'Unit Price', 'Total Price', 'Order Status', 'Order Total',
        'Discount', 'Tax', 'Order Date', 'Delivery Date', 'Notes'
    ]
    ws_details.append(detail_headers)
    
    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws_details[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add order details data
    for order in orders:
        for item in order.order_items:
            row_data = [
                order.id,
                order.order_number,
                order.customer.name,
                order.customer.email,
                order.customer.phone,
                order.customer.location.name,
                order.agency.name,
                order.salesperson.full_name,
                item.product_name or item.product.name,
                item.product.sku,
                item.quantity,
                float(item.unit_price),
                float(item.total_price),
                order.status,
                float(order.total_amount),
                float(order.discount),
                float(order.tax),
                order.order_date.strftime('%Y-%m-%d %H:%M:%S'),
                order.delivery_date.strftime('%Y-%m-%d') if order.delivery_date else '',
                order.notes
            ]
            ws_details.append(row_data)
    
    # Auto-size columns for details
    for column in ws_details.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_details.column_dimensions[column_letter].width = adjusted_width
    
    # Order summary sheet
    ws_summary = wb.create_sheet("Order Summary")
    summary_headers = [
        'Order Number', 'Customer', 'Agency', 'Salesperson', 'Status',
        'Total Amount', 'Order Date', 'Items Count'
    ]
    ws_summary.append(summary_headers)
    
    # Style summary headers
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add summary data
    order_summary = {}
    for order in orders:
        if order.id not in order_summary:
            row_data = [
                order.order_number,
                order.customer.name,
                order.agency.name,
                order.salesperson.full_name,
                order.status,
                float(order.total_amount),
                order.order_date.strftime('%Y-%m-%d %H:%M:%S'),
                len(order.order_items)
            ]
            ws_summary.append(row_data)
            order_summary[order.id] = True
    
    # Auto-size columns for summary
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_purchase_orders_to_excel(purchase_orders):
    """Export purchase orders to Excel file using openpyxl"""
    wb = Workbook()
    
    # PO details sheet
    ws_details = wb.active
    ws_details.title = "PO Details"
    
    # Headers for details
    detail_headers = [
        'PO ID', 'PO Number', 'Supplier', 'Supplier Email', 'Supplier Phone',
        'Agency', 'Product Name', 'Product SKU', 'Quantity Ordered', 'Quantity Received',
        'Unit Cost', 'Total Cost', 'Status', 'PO Total', 'Created Date', 'Notes'
    ]
    ws_details.append(detail_headers)
    
    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws_details[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add PO details data
    for po in purchase_orders:
        for item in po.po_items:
            row_data = [
                po.id,
                po.po_number,
                po.supplier.name,
                po.supplier.email or '',
                po.supplier.phone or '',
                po.agency_ref.name,
                item.product_name or item.product.name,
                item.product.sku,
                item.quantity_ordered,
                item.quantity_received or 0,
                float(item.unit_cost),
                float(item.total_cost),
                po.status,
                float(po.total_amount),
                po.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                po.notes or ''
            ]
            ws_details.append(row_data)
    
    # Auto-size columns for details
    for column in ws_details.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_details.column_dimensions[column_letter].width = adjusted_width
    
    # PO summary sheet
    ws_summary = wb.create_sheet("PO Summary")
    summary_headers = [
        'PO Number', 'Supplier', 'Agency', 'Status',
        'Total Amount', 'Created Date', 'Items Count'
    ]
    ws_summary.append(summary_headers)
    
    # Style summary headers
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add summary data
    po_summary = {}
    for po in purchase_orders:
        if po.id not in po_summary:
            row_data = [
                po.po_number,
                po.supplier.name,
                po.agency_ref.name,
                po.status,
                float(po.total_amount),
                po.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                len(po.po_items)
            ]
            ws_summary.append(row_data)
            po_summary[po.id] = True
    
    # Auto-size columns for summary
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_pos_sales_to_excel(orders):
    """Export POS sales to Excel file using openpyxl"""
    wb = Workbook()
    
    # Sales details sheet
    ws_details = wb.active
    ws_details.title = "Sales Details"
    
    # Headers for details
    detail_headers = [
        'Receipt #', 'Date/Time', 'Customer', 'Customer Phone', 'Agency',
        'Product Name', 'Product SKU', 'Quantity', 'Unit Price', 'Discount %',
        'Line Total', 'Status', 'Sale Total', 'Notes'
    ]
    ws_details.append(detail_headers)
    
    # Style headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws_details[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add sales details data
    for order in orders:
        for item in order.order_items:
            row_data = [
                order.order_number,
                order.order_date.strftime('%Y-%m-%d %H:%M:%S'),
                order.customer.name,
                order.customer.phone or '',
                order.agency.name,
                item.product_name or item.product.name,
                item.product.sku,
                item.quantity,
                float(item.unit_price),
                float(item.discount_percentage or 0),
                float(item.total_price),
                order.status,
                float(order.total_amount),
                order.notes or ''
            ]
            ws_details.append(row_data)
    
    # Auto-size columns for details
    for column in ws_details.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_details.column_dimensions[column_letter].width = adjusted_width
    
    # Sales summary sheet
    ws_summary = wb.create_sheet("Sales Summary")
    summary_headers = [
        'Receipt #', 'Date/Time', 'Customer', 'Agency', 'Status',
        'Items Count', 'Total Amount'
    ]
    ws_summary.append(summary_headers)
    
    # Style summary headers
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add summary data
    for order in orders:
        row_data = [
            order.order_number,
            order.order_date.strftime('%Y-%m-%d %H:%M:%S'),
            order.customer.name,
            order.agency.name,
            order.status,
            len(order.order_items),
            float(order.total_amount)
        ]
        ws_summary.append(row_data)
    
    # Auto-size columns for summary
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_bulk_order_template(order_type='sale'):
    """
    Generate Excel template for bulk order creation
    order_type: 'sale' or 'purchase'
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Order Template"
    
    # Define headers based on order type
    if order_type == 'sale':
        headers = [
            'Order Group ID',
            'Customer Name',
            'Order Date (YYYY-MM-DD)',
            'Product Code',
            'Product Name',
            'UOM',
            'Quantity',
            'Rate',
            'Tax %',
            'Discount %',
            'Payment Type',
            'Delivery Type',
            'Remarks'
        ]
        instructions = [
            'Instructions:',
            '1. Order Group ID: Use same ID for products in one order (e.g., ORD001, ORD002). Leave blank for single product orders.',
            '2. Customer Name: Must match existing customer name exactly',
            '3. Order Date: Format YYYY-MM-DD (e.g., 2024-01-15)',
            '4. Product Code: Product SKU code',
            '5. Payment Type: Cash, Credit, or Credit Sale',
            '6. Delivery Type: Local or Others',
            '7. Mandatory fields: Customer Name, Product Code, Quantity, Rate, Payment Type',
            '8. Skip blank rows - they will be ignored during upload'
        ]
    else:  # purchase
        headers = [
            'Order Group ID',
            'Supplier Name',
            'Order Date (YYYY-MM-DD)',
            'Product Code',
            'Product Name',
            'UOM',
            'Quantity',
            'Rate',
            'Tax %',
            'Discount %',
            'Payment Type',
            'Remarks'
        ]
        instructions = [
            'Instructions:',
            '1. Order Group ID: Use same ID for products in one order (e.g., PO001, PO002). Leave blank for single product orders.',
            '2. Supplier Name: Must match existing supplier name exactly',
            '3. Order Date: Format YYYY-MM-DD (e.g., 2024-01-15)',
            '4. Product Code: Product SKU code',
            '5. Payment Type: Cash, Credit, or Credit Sale',
            '6. Mandatory fields: Supplier Name, Product Code, Quantity, Rate, Payment Type',
            '7. Skip blank rows - they will be ignored during upload'
        ]
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions", 0)
    for idx, instruction in enumerate(instructions, start=1):
        ws_instructions.cell(row=idx, column=1, value=instruction)
    
    # Style instructions
    from openpyxl.styles import Alignment
    for row in ws_instructions.iter_rows(min_row=1, max_row=len(instructions)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws_instructions.column_dimensions['A'].width = 100
    
    # Add headers to template sheet
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Add sample data rows
    if order_type == 'sale':
        sample_rows = [
            ['ORD001', 'ABC Company', '2024-01-15', 'PROD001', 'Sample Product 1', 'pcs', 10, 100, 18, 5, 'Cash', 'Local', 'Sample order'],
            ['ORD001', 'ABC Company', '2024-01-15', 'PROD002', 'Sample Product 2', 'kg', 5, 200, 18, 0, 'Cash', 'Local', 'Sample order'],
            ['', 'XYZ Ltd', '2024-01-16', 'PROD003', 'Sample Product 3', 'ltr', 20, 50, 12, 10, 'Credit', 'Others', 'Single product order'],
        ]
    else:
        sample_rows = [
            ['PO001', 'Supplier ABC', '2024-01-15', 'PROD001', 'Sample Product 1', 'pcs', 100, 80, 18, 0, 'Credit', 'Bulk purchase'],
            ['PO001', 'Supplier ABC', '2024-01-15', 'PROD002', 'Sample Product 2', 'kg', 50, 150, 18, 5, 'Credit', 'Bulk purchase'],
            ['', 'Supplier XYZ', '2024-01-16', 'PROD003', 'Sample Product 3', 'ltr', 200, 40, 12, 0, 'Cash', 'Single product order'],
        ]
    
    for row in sample_rows:
        ws.append(row)
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def process_bulk_orders(file, order_type, agency_id, user_id, user_role):
    """
    Process bulk order upload from Excel file
    order_type: 'sale' or 'purchase'
    Returns: dict with success/error details
    """
    from openpyxl import load_workbook
    from models import Customer, Supplier, Product, Order, OrderItem, PurchaseOrder, PurchaseOrderItem
    from decimal import Decimal
    import uuid
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Group orders by Order Group ID
        order_groups = {}
        single_orders = []
        
        results = {
            'success': [],
            'errors': [],
            'skipped': 0
        }
        
        # Process data rows
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                results['skipped'] += 1
                continue
            
            # Create row dictionary
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = value
            
            # Extract and validate data
            order_group_id = str(row_dict.get('Order Group ID', '')).strip()
            entity_name = str(row_dict.get('Customer Name' if order_type == 'sale' else 'Supplier Name', '')).strip()
            order_date_str = str(row_dict.get('Order Date (YYYY-MM-DD)', '')).strip()
            product_code = str(row_dict.get('Product Code', '')).strip()
            quantity = row_dict.get('Quantity', 0)
            rate = row_dict.get('Rate', 0)
            payment_type = str(row_dict.get('Payment Type', '')).strip()
            
            # Validate mandatory fields
            if not all([entity_name, product_code, quantity, rate, payment_type]):
                results['errors'].append({
                    'row': row_num,
                    'error': 'Missing mandatory fields (Customer/Supplier Name, Product Code, Quantity, Rate, Payment Type)'
                })
                continue
            
            # Parse order date
            try:
                if order_date_str:
                    order_date = datetime.strptime(order_date_str, '%Y-%m-%d')
                else:
                    order_date = datetime.utcnow()
            except ValueError:
                results['errors'].append({
                    'row': row_num,
                    'error': f'Invalid date format: {order_date_str}. Use YYYY-MM-DD'
                })
                continue
            
            # Validate entity (Customer or Supplier)
            if order_type == 'sale':
                from sqlalchemy import func
                entity = Customer.query.filter(
                    func.lower(Customer.name) == entity_name.lower(),
                    Customer.is_active == True
                ).first()
                if not entity:
                    results['errors'].append({
                        'row': row_num,
                        'error': f'Customer not found: {entity_name}'
                    })
                    continue
                # Check agency access
                if user_role != 'super_admin' and entity.location.agency_id != agency_id:
                    results['errors'].append({
                        'row': row_num,
                        'error': f'Customer does not belong to your agency: {entity_name}'
                    })
                    continue
            else:  # purchase
                from sqlalchemy import func
                entity = Supplier.query.filter(
                    func.lower(Supplier.name) == entity_name.lower(),
                    Supplier.is_active == True
                ).first()
                if not entity:
                    results['errors'].append({
                        'row': row_num,
                        'error': f'Supplier not found: {entity_name}'
                    })
                    continue
                # Check agency access
                if user_role != 'super_admin' and entity.agency_id != agency_id:
                    results['errors'].append({
                        'row': row_num,
                        'error': f'Supplier does not belong to your agency: {entity_name}'
                    })
                    continue
            
            # Validate product
            product = Product.query.filter_by(sku=product_code, is_active=True).first()
            if not product:
                results['errors'].append({
                    'row': row_num,
                    'error': f'Product not found: {product_code}'
                })
                continue
            
            # Prepare item data
            item_data = {
                'product': product,
                'product_name': str(row_dict.get('Product Name', product.name)).strip(),
                'uom': str(row_dict.get('UOM', product.uom_ref.short_name if product.uom_ref else 'pcs')).strip(),
                'quantity': float(quantity),
                'rate': float(rate),
                'tax_percentage': float(row_dict.get('Tax %', 0) or 0),
                'discount_percentage': float(row_dict.get('Discount %', 0) or 0),
                'payment_type': payment_type,
                'remarks': str(row_dict.get('Remarks', '')).strip(),
                'row_num': row_num
            }
            
            if order_type == 'sale':
                item_data['delivery_type'] = str(row_dict.get('Delivery Type', 'Local')).strip()
            
            # Group by Order Group ID or create single order
            if order_group_id:
                if order_group_id not in order_groups:
                    order_groups[order_group_id] = {
                        'entity': entity,
                        'entity_name': entity_name,
                        'order_date': order_date,
                        'payment_type': payment_type,
                        'delivery_type': item_data.get('delivery_type', 'Local'),
                        'remarks': item_data['remarks'],
                        'items': []
                    }
                order_groups[order_group_id]['items'].append(item_data)
            else:
                single_orders.append({
                    'entity': entity,
                    'entity_name': entity_name,
                    'order_date': order_date,
                    'payment_type': payment_type,
                    'delivery_type': item_data.get('delivery_type', 'Local'),
                    'remarks': item_data['remarks'],
                    'items': [item_data]
                })
        
        # Create orders from groups
        all_orders = list(order_groups.values()) + single_orders
        
        for order_data in all_orders:
            try:
                if order_type == 'sale':
                    # Create Sales Order
                    order_number = f"ORD-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"
                    
                    order = Order(
                        order_number=order_number,
                        customer_id=order_data['entity'].id,
                        agency_id=order_data['entity'].location.agency_id,
                        salesperson_id=user_id,
                        status='pending',
                        payment_status='pending',
                        order_date=order_data['order_date'],
                        payment_mode=order_data['payment_type'].lower().replace(' ', '_'),
                        order_type=order_data['delivery_type'].lower(),
                        notes=order_data['remarks']
                    )
                    db.session.add(order)
                    db.session.flush()
                    
                    subtotal = Decimal('0')
                    total_tax = Decimal('0')
                    total_discount = Decimal('0')
                    
                    for item_data in order_data['items']:
                        quantity = Decimal(str(item_data['quantity']))
                        unit_price = Decimal(str(item_data['rate']))
                        discount_pct = Decimal(str(item_data['discount_percentage']))
                        tax_pct = Decimal(str(item_data['tax_percentage']))
                        
                        # Calculate line totals
                        line_subtotal = quantity * unit_price
                        line_discount = line_subtotal * (discount_pct / 100)
                        line_after_discount = line_subtotal - line_discount
                        line_tax = line_after_discount * (tax_pct / 100)
                        line_total = line_after_discount + line_tax
                        
                        discounted_price = unit_price * (1 - (discount_pct / 100))
                        
                        order_item = OrderItem(
                            order_id=order.id,
                            product_id=item_data['product'].id,
                            quantity=quantity,
                            uom=item_data['uom'],
                            unit_price=unit_price,
                            mrp_price=item_data['product'].mrp_price or unit_price,
                            discount_percentage=discount_pct,
                            discounted_price=discounted_price,
                            tax_code=f'GST{int(tax_pct)}',
                            tax_rate=tax_pct,
                            tax_amount=line_tax,
                            line_total=line_total,
                            total_price=line_total
                        )
                        db.session.add(order_item)
                        
                        subtotal += line_subtotal
                        total_tax += line_tax
                        total_discount += line_discount
                    
                    order.subtotal_amount = subtotal
                    order.total_tax_amount = total_tax
                    order.discount = total_discount
                    order.total_amount = subtotal - total_discount + total_tax
                    order.total_items_count = len(order_data['items'])
                    
                    db.session.commit()
                    
                    results['success'].append({
                        'order_number': order_number,
                        'entity': order_data['entity_name'],
                        'items_count': len(order_data['items']),
                        'total': float(order.total_amount)
                    })
                    
                else:  # purchase order
                    po_number = f"PO-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"
                    
                    # Determine agency_id for PO
                    po_agency_id = order_data['entity'].agency_id if hasattr(order_data['entity'], 'agency_id') else agency_id
                    
                    purchase_order = PurchaseOrder(
                        po_number=po_number,
                        supplier_id=order_data['entity'].id,
                        agency_id=po_agency_id,
                        created_by=user_id,
                        status='draft',
                        notes=order_data['remarks']
                    )
                    db.session.add(purchase_order)
                    db.session.flush()
                    
                    total_amount = Decimal('0')
                    
                    for item_data in order_data['items']:
                        quantity = int(item_data['quantity'])
                        unit_cost = Decimal(str(item_data['rate']))
                        line_total = quantity * unit_cost
                        
                        po_item = PurchaseOrderItem(
                            po_id=purchase_order.id,
                            product_id=item_data['product'].id,
                            quantity_ordered=quantity,
                            unit_cost=unit_cost,
                            total_cost=line_total
                        )
                        db.session.add(po_item)
                        total_amount += line_total
                    
                    purchase_order.total_amount = total_amount
                    db.session.commit()
                    
                    results['success'].append({
                        'order_number': po_number,
                        'entity': order_data['entity_name'],
                        'items_count': len(order_data['items']),
                        'total': float(purchase_order.total_amount)
                    })
                    
            except Exception as e:
                db.session.rollback()
                results['errors'].append({
                    'row': 'Order Creation',
                    'error': f'Failed to create order for {order_data["entity_name"]}: {str(e)}'
                })
        
        return results
        
    except Exception as e:
        db.session.rollback()
        return {
            'success': [],
            'errors': [{'row': 'File Processing', 'error': str(e)}],
            'skipped': 0
        }