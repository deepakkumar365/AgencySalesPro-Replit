import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

from extensions import db
from models import Order, OrderItem, Customer, Product, Agency, User

def export_orders_to_excel(orders):
    """Exports a list of orders to an Excel file in memory."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Sales Orders'

    # Header
    header = [
        'Order Number', 'Order Date', 'Customer Name', 'Customer Email', 'Salesperson',
        'Agency', 'Status', 'Payment Status', 'Total Items', 'Subtotal', 'Tax', 'Discount', 'Total Amount'
    ]
    sheet.append(header)

    # Data
    for order in orders:
        row = [
            order.order_number,
            order.order_date.strftime('%Y-%m-%d'),
            order.customer.name,
            order.customer.email,
            order.salesperson.get_full_name() if order.salesperson else '',
            order.agency.name if order.agency else '',
            order.status,
            order.payment_status,
            order.total_items_count,
            order.subtotal_amount,
            order.total_tax_amount,
            order.discount,
            order.total_amount
        ]
        sheet.append(row)

    # Styling (optional)
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font

    # Auto-fit columns
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Save to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def generate_bulk_order_template(order_type='sale'):
    """Generates an Excel template for bulk order uploads."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    if order_type == 'sale':
        sheet.title = 'Bulk Sales Order Template'
        # Define header based on required fields for a sales order
        header = [
            'order_date',         # YYYY-MM-DD
            'customer_code',      # Unique Customer Code
            'product_sku',        # Unique Product SKU
            'quantity',           # Number
            'unit_price',         # Price per unit
            'discount_percentage' # 0-100
        ]
        # Sample data
        sample_data = [
            (datetime.now().strftime('%Y-%m-%d'), 'CUST01', 'SKU001', 10, 150.00, 5),
            (datetime.now().strftime('%Y-%m-%d'), 'CUST02', 'SKU002', 5, 200.00, 0)
        ]
    else: # purchase
        sheet.title = 'Bulk Purchase Order Template'
        header = [
            'order_date', 
            'supplier_code', 
            'product_sku', 
            'quantity', 
            'unit_cost'
        ]
        sample_data = [
            (datetime.now().strftime('%Y-%m-%d'), 'SUPP01', 'SKU001', 50, 120.00),
        ]

    # Write header
    sheet.append(header)
    
    # Write sample data
    for row_data in sample_data:
        sheet.append(row_data)

    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Auto-fit columns
    for i, column_cells in enumerate(sheet.columns):
        length = max(len(str(cell.value) or "") for cell in column_cells)
        sheet.column_dimensions[get_column_letter(i + 1)].width = length + 5

    # Save to a BytesIO object
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def process_bulk_orders(file, order_type, agency_id, user_id, user_role):
    """
    Placeholder for the function that processes the uploaded Excel file.
    This would contain the logic to read the Excel file and create/update orders.
    """
    # This is where you would use a library like pandas or openpyxl to read the file
    # and iterate through rows to create orders.
    # For example:
    # import pandas as pd
    # df = pd.read_excel(file)
    # for index, row in df.iterrows():
    #   customer_code = row['customer_code']
    #   product_sku = row['product_sku']
    #   ... etc ...
    #   # Find customer, product, create order and order items
    
    # Returning dummy data for now.
    results = {
        'success': [],
        'errors': [{'row': 3, 'error': 'Example error: Product with SKU "INVALID" not found.'}],
        'skipped': 1
    }
    return results